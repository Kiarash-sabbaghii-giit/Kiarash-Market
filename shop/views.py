import json
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.db import IntegrityError
from datetime import datetime
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import UserInfo, Cart, Order, OrderItem
from .mongo_utils import get_category_info, search_products, get_all_categories_sorted, get_products_by_category, get_db


# ==================== بررسی ادمین ====================
def is_admin(user):
    """بررسی اینکه کاربر ادمین است (فقط خودت)"""
    return user.is_authenticated and user.username == 'kiarash'


# ==================== صفحات اصلی ====================

def home(request):
    return render(request, 'shop/home.html')


def about(request):
    return render(request, 'shop/about.html')


def support(request):
    return render(request, 'shop/support.html')


# ==================== دسته بندی و محصولات ====================

def category_list(request):
    categories = get_all_categories_sorted()
    return render(request, 'shop/categories.html', {'categories': categories})


def category_products(request, category_en):
    category_info = get_category_info(category_en)
    if not category_info:
        return render(request, 'shop/404.html', {'message': 'دسته بندی یافت نشد'})

    products = get_products_by_category(category_en)
    paginator = Paginator(products, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'shop/products.html', {
        'products': page_obj,
        'category': category_info,
        'category_en': category_en,
    })


# ==================== جستجو ====================

def search_suggestions(request):
    query = request.GET.get('q', '')
    if len(query) >= 2:
        results = search_products(query)
        return JsonResponse({'results': results}, safe=False)
    return JsonResponse({'results': []})


def search_page(request):
    query = request.GET.get('q', '')
    products = []
    if query:
        products = search_products(query)

    paginator = Paginator(products, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'shop/search.html', {
        'products': page_obj,
        'query': query,
        'count': len(products)
    })


# ==================== احراز هویت ====================

def register_view(request):
    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        phone = request.POST.get('phone')
        home_phone = request.POST.get('home_phone')
        address = request.POST.get('address')
        plaque = request.POST.get('plaque')
        unit = request.POST.get('unit')
        email = request.POST.get('email')

        if UserInfo.objects.filter(phone=phone).exists():
            messages.error(request, 'این شماره قبلاً ثبت نام کرده!')
            return redirect('register')

        try:
            UserInfo.objects.create(
                first_name=first_name,
                last_name=last_name,
                phone=phone,
                home_phone=home_phone,
                address=address,
                plaque=plaque,
                unit=unit,
                email=email
            )
            messages.success(request, f'سلام {first_name}! ثبت نامت موفق بود.')
            return redirect('login')
        except IntegrityError:
            messages.error(request, 'خطا در ثبت نام')
            return redirect('register')

    return render(request, 'shop/register.html')


def login_view(request):
    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        phone = request.POST.get('phone')

        try:
            user = UserInfo.objects.get(
                first_name=first_name,
                last_name=last_name,
                phone=phone
            )
            request.session['user_id'] = user.id
            request.session['user_name'] = user.first_name
            request.session['user_lastname'] = user.last_name
            request.session['user_phone'] = user.phone
            request.session['user_email'] = user.email or ''

            messages.success(request, f'خوش اومدی {user.first_name}!')
            return redirect('home')
        except UserInfo.DoesNotExist:
            messages.error(request, 'اطلاعات صحیح نیست! ثبت نام کردی؟')
            return redirect('login')

    return render(request, 'shop/login.html')


def logout_view(request):
    if 'user_id' in request.session:
        del request.session['user_id']
        del request.session['user_name']
        del request.session['user_lastname']
        del request.session['user_phone']
        del request.session['user_email']
    messages.success(request, 'با موفقیت خارج شدی!')
    return redirect('home')


# ==================== پروفایل ====================

def profile_view(request):
    if not request.session.get('user_id'):
        messages.warning(request, 'لطفاً ابتدا وارد شوید')
        return redirect('login')

    user = UserInfo.objects.get(id=request.session['user_id'])

    if request.method == 'POST':
        user.first_name = request.POST.get('first_name')
        user.last_name = request.POST.get('last_name')
        user.email = request.POST.get('email')
        user.home_phone = request.POST.get('home_phone')
        user.address = request.POST.get('address')
        user.plaque = request.POST.get('plaque')
        user.unit = request.POST.get('unit')
        user.national_code = request.POST.get('national_code')

        birth_date = request.POST.get('birth_date')
        if birth_date:
            try:
                user.birth_date = datetime.strptime(birth_date, '%Y-%m-%d').date()
            except:
                pass

        user.save()

        request.session['user_name'] = user.first_name
        request.session['user_lastname'] = user.last_name
        request.session['user_email'] = user.email or ''

        messages.success(request, 'مشخصات شما با موفقیت ویرایش شد')
        return redirect('profile')

    return render(request, 'shop/profile.html', {'user': user})


def orders_view(request):
    if not request.session.get('user_id'):
        messages.warning(request, 'لطفاً ابتدا وارد شوید')
        return redirect('login')

    user_id = request.session.get('user_id')
    orders = Order.objects.filter(user_id=user_id).order_by('-created_at')

    return render(request, 'shop/orders.html', {'orders': orders})


# ==================== سبد خرید ====================

def extract_price(price_str):
    if not price_str:
        return 0
    price_str = str(price_str)
    price_str = price_str.replace(',', '').replace('٬', '').replace('تومان', '').replace(' ', '').strip()
    import re
    numbers = re.findall(r'\d+', price_str)
    if numbers:
        full_number = ''.join(numbers)
        try:
            return int(full_number)
        except:
            return 0
    return 0


def cart_add(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        user_id = request.session.get('user_id')

        if not user_id:
            return JsonResponse({'error': 'لطفاً ابتدا وارد شوید'}, status=401)

        product_name = data.get('name')
        product_price = data.get('price')
        product_image = data.get('image')

        # استاندارد سازی قیمت
        if product_price:
            product_price = product_price.replace('٬', ',').strip()
        else:
            product_price = "۰ تومان"

        cart_item, created = Cart.objects.get_or_create(
            user_id=user_id,
            product_name=product_name,
            defaults={
                'product_price': product_price,
                'product_image': product_image,
                'quantity': 1
            }
        )

        if not created:
            cart_item.quantity += 1
            cart_item.save()

        cart_count = Cart.objects.filter(user_id=user_id).count()

        return JsonResponse({
            'success': True,
            'message': 'محصول به سبد خرید اضافه شد',
            'cart_count': cart_count
        })

    return JsonResponse({'error': 'درخواست نامعتبر'}, status=400)


def cart_view(request):
    if not request.session.get('user_id'):
        return redirect('login')

    user_id = request.session.get('user_id')
    cart_items = Cart.objects.filter(user_id=user_id)

    total = 0
    for item in cart_items:
        total += item.get_total_int()

    total_formatted = f"{total:,} تومان" if total > 0 else "۰ تومان"

    return render(request, 'shop/cart.html', {
        'cart_items': cart_items,
        'total': total_formatted,
        'total_int': total
    })


def cart_update(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        user_id = request.session.get('user_id')
        product_name = data.get('product_name')
        action = data.get('action')

        if not user_id:
            return JsonResponse({'error': 'لطفاً وارد شوید'}, status=401)

        try:
            cart_item = Cart.objects.get(user_id=user_id, product_name=product_name)
            price = extract_price(cart_item.product_price)

            if action == 'increase':
                cart_item.quantity += 1
                cart_item.save()
            elif action == 'decrease':
                if cart_item.quantity > 1:
                    cart_item.quantity -= 1
                    cart_item.save()
                else:
                    cart_item.delete()
            elif action == 'remove':
                cart_item.delete()
            else:
                return JsonResponse({'error': 'عملیات نامعتبر'}, status=400)

            cart_items = Cart.objects.filter(user_id=user_id)
            total = 0
            for item in cart_items:
                total += item.get_total_int()

            total_formatted = f"{total:,} تومان" if total > 0 else "۰ تومان"

            if action == 'remove' or (action == 'decrease' and cart_item.quantity == 0):
                return JsonResponse({
                    'success': True,
                    'item_removed': True,
                    'total': total_formatted,
                    'cart_count': cart_items.count()
                })
            else:
                item_total = price * cart_item.quantity
                return JsonResponse({
                    'success': True,
                    'new_quantity': cart_item.quantity,
                    'item_total': f"{item_total:,} تومان",
                    'total': total_formatted,
                    'cart_count': cart_items.count()
                })
        except Cart.DoesNotExist:
            return JsonResponse({'error': 'محصول یافت نشد'}, status=404)

    return JsonResponse({'error': 'درخواست نامعتبر'}, status=400)


def cart_count(request):
    user_id = request.session.get('user_id')
    if user_id:
        count = Cart.objects.filter(user_id=user_id).count()
        return JsonResponse({'count': count})
    return JsonResponse({'count': 0})


# ==================== تسویه حساب و سفارشات ====================

def checkout_view(request):
    if not request.session.get('user_id'):
        return redirect('login')

    user_id = request.session.get('user_id')
    cart_items = Cart.objects.filter(user_id=user_id)

    if not cart_items:
        return redirect('cart')

    user = UserInfo.objects.get(id=user_id)

    total = 0
    for item in cart_items:
        total += item.get_total_int()

    total_formatted = f"{total:,} تومان"

    if request.method == 'POST':
        payment_method = request.POST.get('payment_method', 'cod')

        order = Order.objects.create(
            user_id=user_id,
            user_name=f"{user.first_name} {user.last_name}",
            user_phone=user.phone,
            user_email=user.email or '',
            user_address=request.POST.get('address', user.address),
            total_amount=total_formatted,
            total_amount_int=total,
            status='paid' if payment_method == 'online' else 'pending'
        )

        for item in cart_items:
            OrderItem.objects.create(
                order=order,
                product_name=item.product_name,
                product_price=item.product_price,
                product_image=item.product_image,
                quantity=item.quantity
            )

        cart_items.delete()

        if payment_method == 'online':
            return redirect('online_payment', order_id=order.id)
        else:
            messages.success(request, f'✅ سفارش شما با موفقیت ثبت شد! شماره سفارش: #{order.id}')
            return redirect('orders')

    return render(request, 'shop/checkout.html', {
        'cart_items': cart_items,
        'total': total_formatted,
        'user': user
    })


def payment_view(request, order_id):
    if not request.session.get('user_id'):
        return redirect('login')

    try:
        order = Order.objects.get(id=order_id, user_id=request.session.get('user_id'))
    except Order.DoesNotExist:
        return redirect('cart')

    return render(request, 'shop/payment.html', {'order': order})


def online_payment_view(request, order_id):
    if not request.session.get('user_id'):
        return redirect('login')

    try:
        order = Order.objects.get(id=order_id, user_id=request.session.get('user_id'))
    except Order.DoesNotExist:
        return redirect('cart')

    return render(request, 'shop/online_payment.html', {'order': order})


def process_payment(request, order_id):
    if not request.session.get('user_id'):
        return redirect('login')

    try:
        order = Order.objects.get(id=order_id, user_id=request.session.get('user_id'))
        order.status = 'paid'
        order.save()
        messages.success(request, f'✅ پرداخت سفارش #{order.id} با موفقیت انجام شد!')
    except Order.DoesNotExist:
        pass

    return redirect('orders')


# ==================== پنل مدیریت (ادمین) ====================

@user_passes_test(is_admin)
def admin_categories(request):
    db = get_db()
    categories = list(db.categories.find({}, {'_id': 0}))

    if request.method == 'POST':
        category_fa = request.POST.get('category_fa')
        category_en = request.POST.get('category_en')
        image_url = request.POST.get('image_url')

        db.categories.insert_one({
            'category_fa': category_fa,
            'category_en': category_en,
            'image_url': image_url
        })
        messages.success(request, 'دستهبندی با موفقیت اضافه شد')
        return redirect('admin_categories')

    return render(request, 'shop/admin/categories.html', {'categories': categories})


@user_passes_test(is_admin)
def admin_category_edit(request, category_en):
    db = get_db()
    category = db.categories.find_one({'category_en': category_en}, {'_id': 0})

    if request.method == 'POST':
        db.categories.update_one(
            {'category_en': category_en},
            {'$set': {
                'category_fa': request.POST.get('category_fa'),
                'category_en': request.POST.get('category_en'),
                'image_url': request.POST.get('image_url')
            }}
        )
        messages.success(request, 'دستهبندی با موفقیت ویرایش شد')
        return redirect('admin_categories')

    return render(request, 'shop/admin/category_edit.html', {'category': category})


@user_passes_test(is_admin)
def admin_category_delete(request, category_en):
    db = get_db()
    db.categories.delete_one({'category_en': category_en})
    messages.success(request, 'دستهبندی با موفقیت حذف شد')
    return redirect('admin_categories')


@user_passes_test(is_admin)
def admin_products(request, category_en=None):
    db = get_db()
    categories = list(db.categories.find({}, {'_id': 0}))

    if category_en:
        category_info = db.categories.find_one({'category_en': category_en}, {'_id': 0})
        if category_info:
            products = list(db[category_en].find({}, {'_id': 0}))
        else:
            products = []
            category_info = {'category_fa': category_en, 'image_url': ''}
        return render(request, 'shop/admin/products.html', {
            'products': products,
            'category_en': category_en,
            'categories': categories,
            'category_info': category_info
        })

    return render(request, 'shop/admin/products_list.html', {'categories': categories})


@user_passes_test(is_admin)
def admin_product_add(request, category_en):
    db = get_db()

    if request.method == 'POST':
        product = {
            'name': request.POST.get('name'),
            'price': request.POST.get('price'),
            'image_url': request.POST.get('image_url'),
            'store': request.POST.get('store', ''),
            'stock': int(request.POST.get('stock', 100))
        }

        db[category_en].insert_one(product)
        messages.success(request, 'محصول با موفقیت اضافه شد')
        return redirect('admin_products', category_en=category_en)

    return render(request, 'shop/admin/product_add.html', {'category_en': category_en})


@user_passes_test(is_admin)
def admin_product_edit(request, category_en, product_name):
    db = get_db()
    product = db[category_en].find_one({'name': product_name}, {'_id': 0})

    if request.method == 'POST':
        db[category_en].update_one(
            {'name': product_name},
            {'$set': {
                'name': request.POST.get('name'),
                'price': request.POST.get('price'),
                'image_url': request.POST.get('image_url'),
                'store': request.POST.get('store'),
                'stock': int(request.POST.get('stock', 100))
            }}
        )
        messages.success(request, 'محصول با موفقیت ویرایش شد')
        return redirect('admin_products', category_en=category_en)

    return render(request, 'shop/admin/product_edit.html', {
        'product': product,
        'category_en': category_en
    })


@user_passes_test(is_admin)
def admin_product_delete(request, category_en, product_name):
    db = get_db()
    db[category_en].delete_one({'name': product_name})
    messages.success(request, 'محصول با موفقیت حذف شد')
    return redirect('admin_products', category_en=category_en)


# ==================== صفحه 404 ====================

def custom_404(request, exception):
    return render(request, 'shop/404.html', {}, status=404)


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import jdatetime
from django.utils import timezone


@user_passes_test(is_admin)
def admin_orders_report(request):
    """گزارش سفارشات برای ادمین"""

    # دریافت همه سفارشات به ترتیب جدیدترین
    orders = Order.objects.all().order_by('-created_at')

    # آمار کلی
    total_orders = orders.count()
    total_sales = sum(order.total_amount_int for order in orders)
    paid_orders = orders.filter(status='paid').count()
    pending_orders = orders.filter(status='pending').count()

    # فیلترها
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if status_filter:
        orders = orders.filter(status=status_filter)

    # تبدیل تاریخ شمسی برای نمایش
    for order in orders:
        order.jcreated_at = jdatetime.date.fromgregorian(date=order.created_at).strftime('%Y/%m/%d')
        order.jtime = order.created_at.strftime('%H:%M')

    context = {
        'orders': orders,
        'total_orders': total_orders,
        'total_sales': f"{total_sales:,} تومان",
        'paid_orders': paid_orders,
        'pending_orders': pending_orders,
        'status_filter': status_filter,
        'STATUS_CHOICES': Order.STATUS_CHOICES,
    }

    return render(request, 'shop/admin/orders_report.html', context)


@user_passes_test(is_admin)
def admin_orders_export_excel(request):
    """خروجی Excel از سفارشات"""
    from openpyxl.styles import Font, Alignment, PatternFill

    orders = Order.objects.all().order_by('-created_at')

    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش سفارشات"

    # هدرها
    headers = ['شماره سفارش', 'نام مشتری', 'شماره تماس', 'آدرس', 'مبلغ', 'وضعیت', 'تاریخ ثبت']
    ws.append(headers)

    # استایل هدر (بولد)
    bold_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = bold_font

    # داده‌ها
    for order in orders:
        jdate = jdatetime.date.fromgregorian(date=order.created_at).strftime('%Y/%m/%d')
        ws.append([
            order.id,
            order.user_name,
            order.user_phone,
            order.user_address[:50] if order.user_address else '',
            order.total_amount,
            dict(Order.STATUS_CHOICES).get(order.status, order.status),
            jdate
        ])

    # تنظیم عرض ستون‌ها
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="orders_report.xlsx"'
    wb.save(response)
    return response


@user_passes_test(is_admin)
def admin_orders_export_pdf(request):
    """خروجی PDF از سفارشات"""
    orders = Order.objects.all().order_by('-created_at')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="orders_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []

    styles = getSampleStyleSheet()
    title_style = styles['Title']

    # عنوان
    title = Paragraph("گزارش سفارشات کیارش مارکت", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))

    # جدول داده‌ها
    data = [['شماره', 'مشتری', 'شماره تماس', 'مبلغ', 'وضعیت', 'تاریخ']]

    for order in orders[:50]:  # حداکثر 50 سفارش در PDF
        jdate = jdatetime.date.fromgregorian(date=order.created_at).strftime('%Y/%m/%d')
        data.append([str(order.id),
            order.user_name,
            order.user_phone,
            order.total_amount,
            dict(Order.STATUS_CHOICES).get(order.status, order.status),
            jdate
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)
    return response


@user_passes_test(is_admin)
def order_details_api(request, order_id):
    """API دریافت جزئیات سفارش برای مودال"""
    try:
        order = Order.objects.get(id=order_id)
        items = []
        for item in order.items.all():
            items.append({
                'product_name': item.product_name,
                'product_price': item.product_price,
                'quantity': item.quantity,
                'total': item.get_total()
            })

        return JsonResponse({
            'success': True,
            'order': {
                'id': order.id,
                'user_name': order.user_name,
                'user_phone': order.user_phone,
                'user_address': order.user_address,
                'total_amount': order.total_amount,
                'status': order.status,
                'status_fa': dict(Order.STATUS_CHOICES).get(order.status, order.status),
                'created_at': order.created_at.strftime('%Y/%m/%d %H:%M')
            },
            'items': items
        })
    except Order.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'سفارش یافت نشد'}, status=404)